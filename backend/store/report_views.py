import io
import datetime
from django.db.models import Sum, Count, F, Q, ExpressionWrapper, DecimalField
from django.utils.timezone import make_aware, get_current_timezone
from django.http import HttpResponse, JsonResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import Sale, Product, Inventory, Customer, SaleItem
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

def parse_date(date_str, default=None):
    if not date_str:
        return default
    try:
        dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        return make_aware(dt, timezone=get_current_timezone())
    except ValueError:
        return default

def export_excel(title, headers, data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Metadata
    ws.append([f"Generated Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([]) # Empty row

    # Headers
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = Font(bold=True)

    # Data
    for row in data:
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def export_pdf(title, headers, data, filename):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph(title, styles['Title']))
    elements.append(Paragraph(f"Generated Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Convert all data to string for PDF Table
    table_data = [headers] + [[str(item) if item is not None else "" for item in row] for row in data]
    
    t = Table(table_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(t)
    doc.build(elements)
    return response

class SalesReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date = parse_date(request.GET.get('start_date'))
        end_date = parse_date(request.GET.get('end_date'))
        payment_method = request.GET.get('payment_method')
        export = request.GET.get('export')

        qs = Sale.objects.select_related('customer', 'created_by').prefetch_related('items')
        
        if start_date:
            qs = qs.filter(sale_date__gte=start_date)
        if end_date:
            end_date = end_date.replace(hour=23, minute=59, second=59)
            qs = qs.filter(sale_date__lte=end_date)
        if payment_method:
            qs = qs.filter(payment_method=payment_method)

        qs = qs.order_by('-sale_date')

        data = []
        total_sales_revenue = 0
        total_orders = 0
        total_discount = 0

        for sale in qs:
            items_count = sum(item.quantity for item in sale.items.all())
            customer_name = sale.customer.full_name if sale.customer else "Guest"
            created_by = sale.created_by.username if sale.created_by else "System"
            
            subtotal = float(sale.subtotal)
            discount = float(sale.discount)
            total = float(sale.total_amount)
            
            total_sales_revenue += total
            total_orders += 1
            total_discount += discount

            data.append({
                "sale_id": sale.id,
                "date": sale.sale_date.strftime("%Y-%m-%d %H:%M"),
                "customer": customer_name,
                "items_count": items_count,
                "subtotal": subtotal,
                "discount": discount,
                "total": total,
                "payment_method": sale.get_payment_method_display(),
                "created_by": created_by
            })

        summary = {
            "Total Sales": total_sales_revenue,
            "Total Orders": total_orders,
            "Total Discount": total_discount,
            "Average Order Value": (total_sales_revenue / total_orders) if total_orders > 0 else 0
        }

        if export == 'excel':
            headers = ['Sale ID', 'Date', 'Customer', 'Items', 'Subtotal', 'Discount', 'Total', 'Payment Method', 'Created By']
            rows = [[d['sale_id'], d['date'], d['customer'], d['items_count'], d['subtotal'], d['discount'], d['total'], d['payment_method'], d['created_by']] for d in data]
            return export_excel("Sales Report", headers, rows, "sales_report.xlsx")
        
        if export == 'pdf':
            headers = ['Sale ID', 'Date', 'Customer', 'Items', 'Subtotal', 'Discount', 'Total', 'Payment Method', 'Created By']
            rows = [[d['sale_id'], d['date'], d['customer'], d['items_count'], d['subtotal'], d['discount'], d['total'], d['payment_method'], d['created_by']] for d in data]
            return export_pdf("Sales Report", headers, rows, "sales_report.pdf")

        return Response({"summary": summary, "data": data})

class ProductReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        sort_by = request.GET.get('sort_by', '-qty_sold')
        export = request.GET.get('export')
        start_date = parse_date(request.GET.get('start_date'))
        end_date = parse_date(request.GET.get('end_date'))

        q_filter = Q()
        if start_date:
            q_filter &= Q(sale_items__sale__sale_date__gte=start_date)
        if end_date:
            end_date_time = end_date.replace(hour=23, minute=59, second=59)
            q_filter &= Q(sale_items__sale__sale_date__lte=end_date_time)

        qs = Product.objects.select_related('category').annotate(
            qty_sold=Sum('sale_items__quantity', filter=q_filter, default=0),
            revenue=Sum('sale_items__total_price', filter=q_filter, default=0)
        )

        if sort_by == 'qty_sold':
            qs = qs.order_by('qty_sold')
        elif sort_by == '-qty_sold':
            qs = qs.order_by('-qty_sold')
        elif sort_by == 'revenue':
            qs = qs.order_by('revenue')
        elif sort_by == '-revenue':
            qs = qs.order_by('-revenue')

        data = []
        total_products = 0
        total_units_sold = 0
        total_revenue_overall = 0
        total_profit_overall = 0

        for p in qs:
            cost_price = float(p.cost_price)
            unit_price = float(p.selling_price)
            qty_sold = p.qty_sold
            revenue = float(p.revenue)
            cost = qty_sold * cost_price
            profit = revenue - cost
            
            if p.current_stock <= 0:
                stock_status = "out_of_stock"
            elif p.current_stock <= p.minimum_stock_level:
                stock_status = "low_stock"
            else:
                stock_status = "in_stock"

            total_products += 1
            total_units_sold += qty_sold
            total_revenue_overall += revenue
            total_profit_overall += profit

            data.append({
                "product_name": p.name,
                "sku": p.sku,
                "category": p.category.name if p.category else "N/A",
                "unit_price": unit_price,
                "cost_price": cost_price,
                "current_stock": p.current_stock,
                "qty_sold": qty_sold,
                "revenue": revenue,
                "profit": profit,
                "stock_status": stock_status
            })

        summary = {
            "Total Products": total_products,
            "Units Sold": total_units_sold,
            "Total Revenue": total_revenue_overall,
            "Total Profit": total_profit_overall
        }

        if export == 'excel':
            headers = ['Product Name', 'SKU', 'Category', 'Unit Price', 'Cost Price', 'Current Stock', 'Units Sold', 'Revenue', 'Profit', 'Stock Status']
            rows = [[d['product_name'], d['sku'], d['category'], d['unit_price'], d['cost_price'], d['current_stock'], d['qty_sold'], d['revenue'], d['profit'], d['stock_status'].replace('_', ' ').title()] for d in data]
            return export_excel("Product Sales Report", headers, rows, "product_report.xlsx")
        
        if export == 'pdf':
            headers = ['Product Name', 'SKU', 'Category', 'Unit Price', 'Cost Price', 'Current Stock', 'Units Sold', 'Revenue', 'Profit', 'Stock Status']
            rows = [[d['product_name'], d['sku'], d['category'], d['unit_price'], d['cost_price'], d['current_stock'], d['qty_sold'], d['revenue'], d['profit'], d['stock_status'].replace('_', ' ').title()] for d in data]
            return export_pdf("Product Sales Report", headers, rows, "product_report.pdf")

        return Response({"summary": summary, "data": data})

class InventoryReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        status_filter = request.GET.get('stock_status')
        export = request.GET.get('export')

        qs = Inventory.objects.select_related('product', 'product__category')
        
        data = []
        total_products = 0
        total_stock = 0
        low_stock = 0
        out_of_stock = 0
        total_value = 0

        for inv in qs:
            status = inv.stock_status
            if status_filter and status != status_filter:
                continue
            
            val = inv.quantity * float(inv.product.cost_price)
            
            if status == 'out_of_stock':
                out_of_stock += 1
            elif status == 'low_stock':
                low_stock += 1
                
            total_products += 1
            total_stock += inv.quantity
            total_value += val

            data.append({
                "product": inv.product.name,
                "sku": inv.product.sku,
                "category": inv.product.category.name if inv.product.category else "N/A",
                "current_stock": inv.quantity,
                "min_stock": inv.product.minimum_stock_level,
                "status": status,
                "stock_value": val,
                "last_updated": inv.last_updated.strftime("%Y-%m-%d %H:%M")
            })

        summary = {
            "Total Products": total_products,
            "Total Stock": total_stock,
            "Low Stock": low_stock,
            "Out of Stock": out_of_stock,
            "Total Inventory Value": total_value
        }

        if export == 'excel':
            headers = ['Product', 'SKU', 'Category', 'Current Stock', 'Min Stock', 'Stock Value', 'Status', 'Last Updated']
            rows = [[d['product'], d['sku'], d['category'], d['current_stock'], d['min_stock'], d['stock_value'], d['status'].replace('_', ' ').title(), d['last_updated']] for d in data]
            return export_excel("Inventory Report", headers, rows, "inventory_report.xlsx")
        
        if export == 'pdf':
            headers = ['Product', 'SKU', 'Category', 'Current Stock', 'Min Stock', 'Stock Value', 'Status', 'Last Updated']
            rows = [[d['product'], d['sku'], d['category'], d['current_stock'], d['min_stock'], d['stock_value'], d['status'].replace('_', ' ').title(), d['last_updated']] for d in data]
            return export_pdf("Inventory Report", headers, rows, "inventory_report.pdf")

        return Response({"summary": summary, "data": data})

class CustomerReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        export = request.GET.get('export')
        start_date = parse_date(request.GET.get('start_date'))
        end_date = parse_date(request.GET.get('end_date'))

        q_filter = Q()
        if start_date:
            q_filter &= Q(sales__sale_date__gte=start_date)
        if end_date:
            end_date_time = end_date.replace(hour=23, minute=59, second=59)
            q_filter &= Q(sales__sale_date__lte=end_date_time)

        qs = Customer.objects.annotate(
            num_purchases=Count('sales', filter=q_filter),
            total_spend=Sum('sales__total_amount', filter=q_filter, default=0)
        ).order_by('-total_spend')

        data = []
        total_customers = Customer.objects.count()
        active_customers = 0
        total_revenue = 0

        for c in qs:
            if c.num_purchases > 0:
                active_customers += 1
                
            last_purchase_obj = c.sales.filter(q_filter).order_by('-sale_date').first()
            last_date = last_purchase_obj.sale_date.strftime("%Y-%m-%d") if last_purchase_obj else "Never"
            t_spend = float(c.total_spend)
            total_revenue += t_spend
            
            data.append({
                "customer_name": c.full_name,
                "email": c.email,
                "phone": c.phone,
                "num_purchases": c.num_purchases,
                "total_spend": t_spend,
                "avg_order_value": (t_spend / c.num_purchases) if c.num_purchases > 0 else 0,
                "last_purchase": last_date
            })

        summary = {
            "Total Customers": total_customers,
            "Active Customers": active_customers,
            "Total Revenue": total_revenue,
            "Average Spend": (total_revenue / active_customers) if active_customers > 0 else 0
        }

        if export == 'excel':
            headers = ['Customer Name', 'Email', 'Phone', 'Purchases', 'Total Spend', 'Avg Order Value', 'Last Purchase']
            rows = [[d['customer_name'], d['email'], d['phone'], d['num_purchases'], d['total_spend'], d['avg_order_value'], d['last_purchase']] for d in data]
            return export_excel("Customer Report", headers, rows, "customer_report.xlsx")
        
        if export == 'pdf':
            headers = ['Customer Name', 'Email', 'Phone', 'Purchases', 'Total Spend', 'Avg Order Value', 'Last Purchase']
            rows = [[d['customer_name'], d['email'], d['phone'], d['num_purchases'], d['total_spend'], d['avg_order_value'], d['last_purchase']] for d in data]
            return export_pdf("Customer Report", headers, rows, "customer_report.pdf")

        return Response({"summary": summary, "data": data})

class FinancialReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date = parse_date(request.GET.get('start_date'))
        end_date = parse_date(request.GET.get('end_date'))
        export = request.GET.get('export')

        qs = Sale.objects.all()
        items_qs = SaleItem.objects.all()

        if start_date:
            qs = qs.filter(sale_date__gte=start_date)
            items_qs = items_qs.filter(sale__sale_date__gte=start_date)
        if end_date:
            end_date_time = end_date.replace(hour=23, minute=59, second=59)
            qs = qs.filter(sale_date__lte=end_date_time)
            items_qs = items_qs.filter(sale__sale_date__lte=end_date_time)

        stats = qs.aggregate(
            total_revenue=Sum('total_amount', default=0),
            total_discount=Sum('discount', default=0),
            orders=Count('id')
        )

        cost_stats = items_qs.aggregate(
            total_cost=Sum(F('cost_price') * F('quantity'), output_field=DecimalField(default=0))
        )

        total_revenue = float(stats['total_revenue'])
        total_discount = float(stats['total_discount'])
        orders = stats['orders']
        total_cost = float(cost_stats['total_cost'] or 0)
        gross_profit = total_revenue - total_cost
        avg_order_value = (total_revenue / orders) if orders > 0 else 0

        summary = {
            "Total Revenue": total_revenue,
            "Total Cost": total_cost,
            "Gross Profit": gross_profit,
            "Total Discounts": total_discount,
            "Total Orders": orders,
            "Average Order Value": avg_order_value
        }

        data = [
            {"metric": "Total Revenue", "value": total_revenue},
            {"metric": "Total Cost", "value": total_cost},
            {"metric": "Gross Profit", "value": gross_profit},
            {"metric": "Total Discounts", "value": total_discount},
            {"metric": "Number of Orders", "value": orders},
            {"metric": "Average Order Value", "value": avg_order_value},
        ]

        if export == 'excel':
            headers = ['Metric', 'Value']
            rows = [[d['metric'], d['value']] for d in data]
            return export_excel("Financial Summary", headers, rows, "financial_report.xlsx")
        
        if export == 'pdf':
            headers = ['Metric', 'Value']
            rows = [[d['metric'], d['value']] for d in data]
            return export_pdf("Financial Summary", headers, rows, "financial_report.pdf")

        return Response({"summary": summary, "data": data})
